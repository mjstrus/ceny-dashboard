"""
Unit 6: Exporter
Generowanie raportów i exportu do Excel/CSV.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from typing import Dict, Tuple


def export_to_excel(
    df: pd.DataFrame,
    segment_summary: Dict,
    revenue_impact: Dict,
    filename: str = "plan_wdrazania.xlsx"
) -> BytesIO:
    """
    Eksportuj raport do Excela z trzema arkuszami.
    
    Args:
        df: DataFrame z danymi i kalkulacjami
        segment_summary: summary per segment
        revenue_impact: agregaty przychodu
        filename: nazwa pliku
    
    Returns:
        BytesIO buffer z Excelem
    """
    
    # Create ExcelWriter
    buffer = BytesIO()
    excel_writer = pd.ExcelWriter(buffer, engine='openpyxl')
    
    # =====================================================================
    # ARKUSZ 1: CLIENTS (szczegóły każdego klienta)
    # =====================================================================
    
    df_export = df[[
        'ID', 'Nazwa', 'Typ_Pełny', 
        'Cena_Bazowa', 'Cena_Nowa', 'Doc_Avg',
        'Przychod_Stary_Mc', 'Przychod_Nowy_Mc', 'Wplyw_Pln',
        'Wzrost_Bazowy_Pct', 'Anulowanie_Rabatu_Pct', 'Wplyw_Pct',
        'Segment', 'Zagroženie', 'Czy_Rabat'
    ]].copy()
    
    # Rename
    df_export.columns = [
        'ID', 'Nazwa', 'Typ Umowy',
        'Cena Bazowa Stara', 'Cena Bazowa Nowa', 'Śr. Dokumentów/mc',
        'Przychód Stary/mc', 'Przychód Nowy/mc', 'Wpływ PLN',
        'Wzrost %', 'Rabat ∅ %', 'Wpływ %',
        'Segment', 'Zagrożenie', 'Ma Rabat'
    ]
    
    # Sortuj po wpływie (descending)
    df_export = df_export.sort_values('Wpływ %', ascending=False, key=lambda x: pd.to_numeric(x, errors='coerce'))
    
    df_export.to_excel(excel_writer, sheet_name='Clients', index=False)
    
    # =====================================================================
    # ARKUSZ 2: SUMMARY (podsumowanie agregowane)
    # =====================================================================
    
    summary_data = {
        'Metrika': [
            'Liczba Klientów',
            'Przychód Dzisiaj (mc)',
            'Przychód Docelowy (mc)',
            'Wpływ Razem PLN (mc)',
            'Wpływ Razem %',
            'Wpływ Roczny PLN',
            'Przychód Roczny Dzisiaj',
            'Przychód Roczny Docelowy'
        ],
        'Wartość': [
            len(df),
            f"${revenue_impact['revenue_old_monthly']:,.2f}",
            f"${revenue_impact['revenue_new_monthly']:,.2f}",
            f"${revenue_impact['impact_pln_monthly']:,.2f}",
            f"{revenue_impact['impact_pct']:.2f}%",
            f"${revenue_impact['impact_pln_annual']:,.2f}",
            f"${revenue_impact['revenue_old_annual']:,.2f}",
            f"${revenue_impact['revenue_new_annual']:,.2f}"
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(excel_writer, sheet_name='Summary', index=False)
    
    # =====================================================================
    # ARKUSZ 3: BY SEGMENT (podsumowanie per segment)
    # =====================================================================
    
    segment_data = []
    for segment, info in segment_summary.items():
        segment_data.append({
            'Segment': segment,
            'Liczba Klientów': info['count'],
            '% Ogółem': f"{info['count_pct']:.1f}%",
            'Śr. Wpływ %': f"{info['avg_impact_pct']:.1f}%",
            'Razem Wpływ PLN': f"${info['total_impact_pln']:,.2f}",
            'Śr. Przychód Stary': f"${info['avg_income_old']:,.2f}",
            'Śr. Przychód Nowy': f"${info['avg_income_new']:,.2f}"
        })
    
    df_by_segment = pd.DataFrame(segment_data)
    df_by_segment.to_excel(excel_writer, sheet_name='By_Segment', index=False)
    
    # =====================================================================
    # FORMATOWANIE
    # =====================================================================
    
    workbook = excel_writer.book
    
    # Navy header + Gold accents (Abacus style)
    navy_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    gold_fill = PatternFill(start_color='E8A000', end_color='E8A000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True, size=12)
    
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Headers
    for ws in workbook.sheetnames:
        sheet = workbook[ws]
        for cell in sheet[1]:
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = header_border
        
        # Auto-resize columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = max_length + 2
    
    # Format Clients sheet — koloruj segmenty
    clients_sheet = workbook['Clients']
    segment_col = None
    
    # Znajdź kolumnę Segment
    for col_num, cell in enumerate(clients_sheet[1], 1):
        if cell.value == 'Segment':
            segment_col = col_num
            break
    
    if segment_col:
        for row_num, row in enumerate(clients_sheet.iter_rows(min_row=2), start=2):
            segment = row[segment_col - 1].value
            
            if '🟢' in str(segment):
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif '🟡' in str(segment):
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif '🔴' in str(segment):
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            else:
                continue
            
            for cell in row:
                cell.fill = fill
    
    excel_writer.close()
    buffer.seek(0)
    return buffer


def export_to_csv(df: pd.DataFrame, filename: str = "plan_wdrazania.csv") -> str:
    """
    Eksportuj do CSV.
    """
    
    df_export = df[[
        'ID', 'Nazwa', 'Typ_Pełny',
        'Cena_Bazowa', 'Cena_Nowa', 'Doc_Avg',
        'Przychod_Stary_Mc', 'Przychod_Nowy_Mc', 'Wplyw_Pln',
        'Wzrost_Bazowy_Pct', 'Anulowanie_Rabatu_Pct', 'Wplyw_Pct',
        'Segment', 'Zagroženie'
    ]].copy()
    
    df_export.columns = [
        'ID', 'Nazwa', 'Typ Umowy',
        'Cena Bazowa Stara', 'Cena Bazowa Nowa', 'Śr. Dokumentów/mc',
        'Przychód Stary/mc', 'Przychód Nowy/mc', 'Wpływ PLN',
        'Wzrost %', 'Rabat ∅ %', 'Wpływ %',
        'Segment', 'Zagrożenie'
    ]
    
    df_export = df_export.sort_values('Wpływ %', ascending=False, key=lambda x: pd.to_numeric(x, errors='coerce'))
    
    return df_export.to_csv(index=False)


def get_markdown_report(segment_summary: Dict, revenue_impact: Dict, threat_count: int) -> str:
    """
    Generuj raport tekstowy w Markdown.
    """
    
    report = f"""
# Plan Wdrażania Ujednolicenia Cen

## Podsumowanie Całościowe

- **Liczba klientów:** {len([s for s in segment_summary.values()])}
- **Przychód dzisiaj:** ${revenue_impact['revenue_old_monthly']:,.2f}/mc (${revenue_impact['revenue_old_annual']:,.2f}/rok)
- **Przychód docelowo:** ${revenue_impact['revenue_new_monthly']:,.2f}/mc (${revenue_impact['revenue_new_annual']:,.2f}/rok)
- **Wzrost przychodu:** +${revenue_impact['impact_pln_monthly']:,.2f}/mc (+{revenue_impact['impact_pct']:.1f}%)
- **Zagrożeni klienci:** {threat_count} (> 20% wzrost)

## Rozkład per Segment

| Segment | Liczba | % Ogółem | Śr. Wpływ | Razem Wpływ |
|---------|--------|---------|----------|------------|
"""
    
    for segment, info in segment_summary.items():
        report += f"| {segment} | {info['count']} | {info['count_pct']:.1f}% | {info['avg_impact_pct']:.1f}% | ${info['total_impact_pln']:,.0f} |\n"
    
    report += """
## Rekomendacje

### 🟢 Zieloni (wzrost ≤ 10%)
- Łatwa komunikacja
- Wysłać standardową notę o zmianach
- Brak negocjacji

### 🟡 Żółci (wzrost 10-20%)
- Wymaga komunikacji
- Wyjaśnić komponenty (wzrost + rabat)
- Być dostępnym do pytań

### 🔴 Czerwoni (wzrost > 20%)
- **PRIORYTET**: wysokie ryzyko rezygnacji
- Indywidualny kontakt
- Rozważyć alternatywy (np. inny typ umowy)
- Negocjacje na wypadek

## Timeline

- **Fala 1:** Zieloni (komunikacja)
- **Fala 2:** Żółci (z czasem na pytania)
- **Fala 3:** Czerwoni (z ofertą alternatywną)
"""
    
    return report
